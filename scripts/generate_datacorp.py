"""Generate the DataCorp s.r.o. synthetic HR dataset.

Produces seven files in notebooks/, all deterministic given SEED:
- datacorp.csv               — main HR table (~1000 rows, 15 columns, planted dirt)
- datacorp_reviews.csv       — performance reviews (~150 rows, Czech free-text)
- datacorp_exit_interviews.csv — exit interviews (~70 rows, Czech free-text with English fragments)
- datacorp_salary_history.csv  — salary changes (~3700 rows, mixed date/decimal formats)
- datacorp_org_chart.csv       — manager hierarchy (~1200 rows, planted cycles and orphans)
- datacorp_tickets.csv         — IT/HR ticket export (~5000 rows, inconsistent taxonomy)
- datacorp_payroll_q3.xlsx     — finance payroll (Excel, schema drift, EUR landmines)

Usage:
    python scripts/generate_datacorp.py
    # or:
    uv run scripts/generate_datacorp.py
"""

from pathlib import Path

import numpy as np
import pandas as pd

SEED = 42
RNG = np.random.default_rng(SEED)
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "notebooks"

# ── Czech name banks ──────────────────────────────────────────────

MALE_FIRST = [
    "Jan", "Petr", "Tomáš", "Martin", "Jakub", "David", "Lukáš", "Filip",
    "Ondřej", "Michal", "Adam", "Vojtěch", "Daniel", "Marek", "Pavel",
    "Josef", "Matěj", "Dominik", "Radek", "Karel",
]
FEMALE_FIRST = [
    "Jana", "Eva", "Anna", "Lucie", "Tereza", "Kateřina", "Petra", "Marie",
    "Martina", "Lenka", "Monika", "Veronika", "Barbora", "Hana", "Markéta",
    "Klára", "Alena", "Simona", "Nikola", "Kristýna",
]
# Surname roots — male form; female form adds "ová" for adjectival surnames
SURNAME_PAIRS = [
    ("Novák", "Nováková"), ("Svoboda", "Svobodová"), ("Novotný", "Novotná"),
    ("Dvořák", "Dvořáková"), ("Černý", "Černá"), ("Procházka", "Procházková"),
    ("Kučera", "Kučerová"), ("Veselý", "Veselá"), ("Horák", "Horáková"),
    ("Němec", "Němcová"), ("Pokorný", "Pokorná"), ("Marek", "Marková"),
    ("Pospíšil", "Pospíšilová"), ("Hájek", "Hájková"), ("Jelínek", "Jelínková"),
    ("Král", "Králová"), ("Růžička", "Růžičková"), ("Beneš", "Benešová"),
    ("Fiala", "Fialová"), ("Sedláček", "Sedláčková"), ("Doležal", "Doležalová"),
    ("Zeman", "Zemanová"), ("Kolář", "Kolářová"), ("Navrátil", "Navrátilová"),
    ("Čermák", "Čermáková"), ("Vaněk", "Vaňková"), ("Urban", "Urbanová"),
    ("Blažek", "Blažková"), ("Kříž", "Křížová"), ("Kopecký", "Kopecká"),
    ("Konečný", "Konečná"), ("Malý", "Malá"), ("Holub", "Holubová"),
    ("Čech", "Čechová"), ("Štěpánek", "Štěpánková"),
]

CITIES = [
    "Praha", "Praha", "Praha", "Praha",  # heavily weighted
    "Brno", "Brno",
    "Ostrava", "Plzeň", "Liberec", "Olomouc",
    "České Budějovice", "Hradec Králové", "Pardubice", "Zlín",
]

# ── Department profiles ──────────────────────────────────────────

DEPARTMENTS = {
    "Vývoj": {
        "headcount": 200,
        "base_salary": 65_000,
        "salary_std": 12_000,
        "avg_tenure_years": 4.0,
        "tenure_std": 2.0,
        "edu_weights": {"SŠ": 0.10, "VŠ Bc.": 0.10, "VŠ Mgr.": 0.60, "VŠ PhD.": 0.20},
        "positions": [
            "Software Developer", "QA Engineer", "Tech Lead", "DevOps Engineer",
            "Data Engineer",
        ],
    },
    "Marketing": {
        "headcount": 140,
        "base_salary": 45_000,
        "salary_std": 8_000,
        "avg_tenure_years": 3.0,
        "tenure_std": 1.5,
        "edu_weights": {"SŠ": 0.15, "VŠ Bc.": 0.25, "VŠ Mgr.": 0.50, "VŠ PhD.": 0.10},
        "positions": [
            "Marketingový specialista", "Content Manager", "SEO Specialist",
            "Brand Manager",
        ],
    },
    "Obchod": {
        "headcount": 185,
        "base_salary": 50_000,
        "salary_std": 10_000,
        "avg_tenure_years": 3.5,
        "tenure_std": 1.8,
        "edu_weights": {"SŠ": 0.30, "VŠ Bc.": 0.20, "VŠ Mgr.": 0.40, "VŠ PhD.": 0.10},
        "positions": [
            "Obchodní zástupce", "Account Manager", "Sales Manager",
            "Business Development",
        ],
    },
    "HR": {
        "headcount": 100,
        "base_salary": 42_000,
        "salary_std": 7_000,
        "avg_tenure_years": 3.0,
        "tenure_std": 1.5,
        "edu_weights": {"SŠ": 0.15, "VŠ Bc.": 0.20, "VŠ Mgr.": 0.55, "VŠ PhD.": 0.10},
        "positions": ["HR Specialist", "Recruiter", "HR Manager", "Payroll Specialist"],
    },
    "Finance": {
        "headcount": 140,
        "base_salary": 55_000,
        "salary_std": 10_000,
        "avg_tenure_years": 4.5,
        "tenure_std": 2.0,
        "edu_weights": {"SŠ": 0.10, "VŠ Bc.": 0.15, "VŠ Mgr.": 0.55, "VŠ PhD.": 0.20},
        "positions": ["Účetní", "Finanční analytik", "Controller", "Auditor"],
    },
    "Podpora": {
        "headcount": 250,
        "base_salary": 35_000,
        "salary_std": 5_000,
        "avg_tenure_years": 2.0,
        "tenure_std": 1.0,
        "edu_weights": {"SŠ": 0.45, "VŠ Bc.": 0.20, "VŠ Mgr.": 0.30, "VŠ PhD.": 0.05},
        "positions": [
            "Zákaznická podpora", "Help Desk Specialist", "Support Team Lead",
            "Technická podpora",
        ],
    },
}


def build_main_df() -> pd.DataFrame:
    """Generate the clean base employee dataset with planted analytical signals."""
    rows = []
    employee_id = 0

    for dept_name, profile in DEPARTMENTS.items():
        for _ in range(profile["headcount"]):
            # Gender
            is_female = RNG.random() < 0.45
            pohlavi = "Ž" if is_female else "M"

            # Name
            if is_female:
                jmeno = RNG.choice(FEMALE_FIRST)
                _, prijmeni = SURNAME_PAIRS[RNG.integers(len(SURNAME_PAIRS))]
            else:
                jmeno = RNG.choice(MALE_FIRST)
                prijmeni, _ = SURNAME_PAIRS[RNG.integers(len(SURNAME_PAIRS))]

            # Tenure (years) — department-specific
            tenure = max(
                0.1,
                RNG.normal(profile["avg_tenure_years"], profile["tenure_std"]),
            )

            # Age — correlated with tenure
            min_age = 22
            age = int(min_age + tenure + RNG.normal(5, 4))
            age = max(22, min(62, age))

            # Hire date
            from datetime import date, timedelta

            hire_date = date(2025, 6, 1) - timedelta(days=int(tenure * 365))
            # Clamp to reasonable range
            if hire_date < date(2015, 1, 1):
                hire_date = date(2015, 1, 1) + timedelta(days=int(RNG.integers(0, 180)))

            # Education — department-weighted
            edu_labels = list(profile["edu_weights"].keys())
            edu_probs = list(profile["edu_weights"].values())
            vzdelani = RNG.choice(edu_labels, p=edu_probs)

            # Salary — base + seniority + education bonus + noise
            salary = profile["base_salary"]
            salary += tenure * 1_500  # seniority bonus
            if vzdelani == "VŠ Mgr.":
                salary += RNG.normal(8_000, 2_000)
            elif vzdelani == "VŠ PhD.":
                salary += RNG.normal(12_000, 3_000)

            # Spurious signal: Marketing gets slight tenure-dependent bump
            # so that raw gap vs Obchod exists but disappears when controlling
            if dept_name == "Obchod":
                salary += tenure * 800  # extra tenure bonus in sales

            salary += RNG.normal(0, profile["salary_std"])
            salary = max(25_000, round(salary / 500) * 500)  # round to 500

            # SŠ salespeople exceptions — a few earn a lot
            if dept_name == "Obchod" and vzdelani == "SŠ" and RNG.random() < 0.15:
                salary = RNG.integers(70_000, 95_000)
                salary = round(salary / 500) * 500

            # Performance — correlated with tenure
            perf_score = 3.0 + 0.15 * tenure + RNG.normal(0, 0.8)
            perf_score = int(np.clip(round(perf_score), 1, 5))

            pozice = RNG.choice(profile["positions"])
            mesto = RNG.choice(CITIES)
            typ_uvazku = RNG.choice(
                ["Plný", "Částečný", "DPP"], p=[0.85, 0.10, 0.05]
            )

            # Email
            email = (
                f"{jmeno.lower()}.{prijmeni.lower()}@datacorp.cz"
                .replace("á", "a").replace("č", "c").replace("ď", "d")
                .replace("é", "e").replace("ě", "e").replace("í", "i")
                .replace("ň", "n").replace("ó", "o").replace("ř", "r")
                .replace("š", "s").replace("ť", "t").replace("ú", "u")
                .replace("ů", "u").replace("ý", "y").replace("ž", "z")
            )

            rows.append({
                "employee_id": employee_id,
                "jmeno": jmeno,
                "prijmeni": prijmeni,
                "email": email,
                "oddeleni": dept_name,
                "pozice": pozice,
                "datum_nastupu": hire_date,
                "plat": salary,
                "hodnoceni_vykonu": perf_score,
                "vzdelani": vzdelani,
                "mesto": mesto,
                "vek": age,
                "pohlavi": pohlavi,
                "typ_uvazku": typ_uvazku,
                "telefon": f"+420 {RNG.integers(600, 800)} {RNG.integers(100, 1000):03d} {RNG.integers(100, 1000):03d}",
            })
            employee_id += 1

    return pd.DataFrame(rows)


def tag_departures(df: pd.DataFrame) -> pd.DataFrame:
    """Tag ~15 employees as departed.

    Departed employees stay in the in-memory universe (side-tables can
    reference them) but are filtered out of the persisted main CSV.
    This creates natural survivorship bias for downstream analysis.
    """
    df = df.copy()
    df["_departed"] = False
    departed_idx = RNG.choice(df.index, size=15, replace=False)
    df.loc[departed_idx, "_departed"] = True
    return df


def build_salary_history(universe: pd.DataFrame) -> pd.DataFrame:
    """Generate 1–6 salary events per employee, chronological per employee.

    Operates on the full universe (including departed employees) so
    departed people retain salary history that is absent from the main CSV.
    """
    rows = []
    for _, emp in universe.iterrows():
        n_events = int(RNG.integers(1, 7))
        try:
            hire = pd.to_datetime(emp["datum_nastupu"])
        except (ValueError, TypeError):
            hire = pd.Timestamp("2020-01-01")
        current_salary = int(emp["plat"]) if pd.notna(emp.get("plat")) else 50_000
        # Reconstruct an earlier starting salary, then walk forward to current_salary.
        # Each event dates between hire and the Q1 2026 snapshot cutoff.
        cutoff = pd.Timestamp("2026-04-01")
        span_days = max(30, (cutoff - hire).days)
        dates = sorted([
            hire + pd.Timedelta(days=int(RNG.integers(30, max(31, span_days))))
            for _ in range(n_events)
        ])
        prev = current_salary - int(sum(RNG.integers(2_000, 8_000) for _ in range(n_events)))
        for d in dates:
            raise_amt = int(RNG.integers(2_000, 8_000))
            new_salary = prev + raise_amt
            duvod = RNG.choice(
                ["raise", "promotion", "correction", "acquisition_harmonization"],
                p=[0.65, 0.20, 0.10, 0.05],
            )
            rows.append({
                "employee_id": int(emp["employee_id"]),
                "datum_zmeny": d.strftime("%Y-%m-%d"),
                "plat_pred": prev,
                "plat_po": new_salary,
                "duvod": duvod,
            })
            prev = new_salary
    df = pd.DataFrame(rows)
    return df.sample(frac=1, random_state=SEED).reset_index(drop=True)


def apply_dirt_salary_history(df: pd.DataFrame, main_df: pd.DataFrame) -> pd.DataFrame:
    """Plant dirt in salary history. `main_df` is the post-dirt main CSV.

    Conflict rates (per the spec) target ratios over employees present
    in both the salary-history and the main CSV.
    """
    df = df.copy()

    # 12. Date format soup
    n = len(df)
    rest_idx = df.index.tolist()
    czech_idx = RNG.choice(rest_idx, size=int(n * 0.30), replace=False)
    remaining = [i for i in rest_idx if i not in set(czech_idx)]
    us_idx = RNG.choice(remaining, size=int(n * 0.15), replace=False)
    remaining = [i for i in remaining if i not in set(us_idx)]
    year_only_idx = RNG.choice(remaining, size=int(n * 0.03), replace=False)
    for i in czech_idx:
        d = pd.to_datetime(df.at[i, "datum_zmeny"])
        df.at[i, "datum_zmeny"] = d.strftime("%d.%m.%Y")
    for i in us_idx:
        d = pd.to_datetime(df.at[i, "datum_zmeny"])
        df.at[i, "datum_zmeny"] = d.strftime("%m/%d/%Y")
    for i in year_only_idx:
        df.at[i, "datum_zmeny"] = str(pd.to_datetime(df.at[i, "datum_zmeny"]).year)

    # 11. Decimal separator mixing in plat_po
    decimal_idx = RNG.choice(df.index, size=int(n * 0.10), replace=False)
    nbsp = " "  # non-breaking space U+00A0
    for i in decimal_idx:
        v = float(df.at[i, "plat_po"])
        df.at[i, "plat_po"] = f"{int(v // 1000)}{nbsp}{int(v) % 1000:03d},{int((v % 1) * 100):02d}"

    # Drive conflicts with main `plat` (10% off-by-small)
    common_ids = set(df["employee_id"]) & set(main_df["employee_id"])
    common_list = list(common_ids)
    RNG.shuffle(common_list)
    off_small = common_list[: int(len(common_list) * 0.10)]
    for emp_id in off_small:
        rows = df[df["employee_id"] == emp_id]
        if len(rows) == 0:
            continue
        latest_idx = rows.index[-1]
        try:
            val = float(str(df.at[latest_idx, "plat_po"]).replace(nbsp, "").replace(",", "."))
            df.at[latest_idx, "plat_po"] = int(val + RNG.integers(1_000, 5_000))
        except (ValueError, TypeError):
            pass

    # 17. Salary history out of order — pick 2 employees, add a backdated correction
    out_of_order_emps = RNG.choice(common_list, size=2, replace=False)
    for emp_id in out_of_order_emps:
        rows = df[df["employee_id"] == emp_id].sort_values("datum_zmeny")
        if len(rows) < 2:
            continue
        backdate = pd.to_datetime(rows.iloc[0]["datum_zmeny"], errors="coerce")
        if pd.isna(backdate):
            continue
        backdate = backdate - pd.Timedelta(days=30)
        df = pd.concat([df, pd.DataFrame([{
            "employee_id": emp_id,
            "datum_zmeny": backdate.strftime("%Y-%m-%d"),
            "plat_pred": int(rows.iloc[0]["plat_pred"]),
            "plat_po": int(float(rows.iloc[0]["plat_pred"]) - 2000),
            "duvod": "correction",
        }])], ignore_index=True)

    # 14. Silent ID collision — reassign 3 of donor's history rows to victim's ID
    if len(common_list) >= 2:
        donor, victim = common_list[0], common_list[1]
        donor_rows = df[df["employee_id"] == donor].head(3).index
        df.loc[donor_rows, "employee_id"] = victim

    return df.reset_index(drop=True)


def apply_dirt_main(df: pd.DataFrame) -> pd.DataFrame:
    """Apply pre-planted data quality issues."""
    df = df.copy()
    n = len(df)

    # 1. Inconsistent oddeleni values
    marketing_mask = df["oddeleni"] == "Marketing"
    marketing_idx = df[marketing_mask].index.tolist()
    if len(marketing_idx) > 20:
        lowercase_idx = RNG.choice(marketing_idx, size=15, replace=False)
        mktg_idx = RNG.choice(
            [i for i in marketing_idx if i not in lowercase_idx], size=5, replace=False
        )
        df.loc[lowercase_idx, "oddeleni"] = "marketing"
        df.loc[mktg_idx, "oddeleni"] = "Mktg"

    # 2. Missing plat — ~40 total, ~25 in Podpora
    podpora_idx = df[df["oddeleni"] == "Podpora"].index.tolist()
    other_idx = df[df["oddeleni"] != "Podpora"].index.tolist()
    missing_podpora = RNG.choice(podpora_idx, size=min(25, len(podpora_idx)), replace=False)
    missing_other = RNG.choice(other_idx, size=min(15, len(other_idx)), replace=False)
    df.loc[missing_podpora, "plat"] = np.nan
    df.loc[missing_other, "plat"] = np.nan

    # 3. Mixed date formats — ~100 rows get DD.MM.YYYY, rest YYYY-MM-DD
    format_idx = RNG.choice(df.index, size=100, replace=False)
    df["datum_nastupu"] = df["datum_nastupu"].apply(
        lambda d: d.strftime("%Y-%m-%d") if not pd.isna(d) else ""
    )
    for idx in format_idx:
        try:
            from datetime import datetime

            d = datetime.strptime(df.at[idx, "datum_nastupu"], "%Y-%m-%d")
            df.at[idx, "datum_nastupu"] = d.strftime("%d.%m.%Y")
        except (ValueError, TypeError):
            pass

    # 4. Salary outliers — 4 extreme values in low-salary departments
    outlier_depts = ["Podpora", "HR"]
    for dept in outlier_depts:
        dept_idx = df[(df["oddeleni"] == dept) & df["plat"].notna()].index.tolist()
        if len(dept_idx) >= 2:
            outlier_idx = RNG.choice(dept_idx, size=2, replace=False)
            df.loc[outlier_idx, "plat"] = RNG.integers(200_000, 260_000, size=2)

    # 5. Duplicate rows — 10 exact copies
    dup_idx = RNG.choice(df.index, size=10, replace=False)
    duplicates = df.loc[dup_idx].copy()
    df = pd.concat([df, duplicates], ignore_index=True)

    # 6. Invalid hodnoceni_vykonu — 5-8 values of 0 or 6
    n_invalid = RNG.integers(5, 9)
    invalid_idx = RNG.choice(df.index, size=n_invalid, replace=False)
    df.loc[invalid_idx, "hodnoceni_vykonu"] = RNG.choice([0, 6], size=n_invalid)

    # 7. Phone-number formats and invalids
    phone_idx = df.index.tolist()
    n_alt = 200  # split across alt formats
    alt_idx = RNG.choice(phone_idx, size=n_alt, replace=False)
    half = n_alt // 2
    for i in alt_idx[:half]:
        df.at[i, "telefon"] = df.at[i, "telefon"].replace("+420 ", "").replace(" ", "")
    for i in alt_idx[half:]:
        df.at[i, "telefon"] = df.at[i, "telefon"].replace("+420 ", "")
    missing_phone = RNG.choice(phone_idx, size=20, replace=False)
    df.loc[missing_phone, "telefon"] = np.nan
    invalid_phone = RNG.choice(phone_idx, size=5, replace=False)
    df.loc[invalid_phone, "telefon"] = RNG.choice(["???", "viz Slack"], size=5)

    # 8. Email / name mismatches (rehires reusing employee_id)
    rehire_idx = RNG.choice(df.index, size=15, replace=False)
    ghost_emails = [
        "petra.svobodova@datacorp.cz", "tomas.novak@datacorp.cz",
        "jana.kralova@datacorp.cz", "marek.benes@datacorp.cz",
        "lucie.dvorakova@datacorp.cz",
    ]
    for k, i in enumerate(rehire_idx):
        df.at[i, "email"] = ghost_emails[k % len(ghost_emails)]

    # 9. Whitespace and embedded newlines in name fields
    ws_idx = RNG.choice(df.index, size=30, replace=False)
    for k, i in enumerate(ws_idx):
        col = "jmeno" if k % 2 == 0 else "prijmeni"
        original = df.at[i, col]
        if k % 3 == 0:
            df.at[i, col] = f" {original}"
        elif k % 3 == 1:
            df.at[i, col] = f"{original} "
        else:
            df.at[i, col] = f" {original} "

    # 10. Encoding mojibake — 3 rows with mangled diacritics
    mojibake_idx = RNG.choice(df.index, size=3, replace=False)
    mojibake_map = str.maketrans({"Č": "Ä", "č": "Ä", "ž": "Å¾", "š": "Å¡"})
    for i in mojibake_idx:
        df.at[i, "prijmeni"] = str(df.at[i, "prijmeni"]).translate(mojibake_map)

    # 16. Future hire dates (typo 2017 -> 2027)
    future_idx = RNG.choice(df.index, size=5, replace=False)
    for i in future_idx:
        existing = str(df.at[i, "datum_nastupu"])
        replaced = existing.replace("2017", "2027").replace("2018", "2028")
        if replaced == existing:
            # No 2017/2018 in the original; force a future date.
            df.at[i, "datum_nastupu"] = "2027-06-15"
        else:
            df.at[i, "datum_nastupu"] = replaced

    # 15. Diacritic-folded duplicate identities — add 2 paired rows
    if "Černý" in df["prijmeni"].astype(str).values:
        cerny_rows = df[df["prijmeni"].astype(str) == "Černý"].head(1).copy()
        if len(cerny_rows):
            folded = cerny_rows.copy()
            folded["prijmeni"] = "Cerny"
            folded["jmeno"] = folded["jmeno"].astype(str).str.replace("á", "a").str.replace("í", "i")
            folded["employee_id"] = int(df["employee_id"].astype(int).max()) + 5000
            folded["email"] = folded["email"].astype(str).str.replace("ý", "y").str.replace("á", "a")
            df = pd.concat([df, folded], ignore_index=True)

    # Shuffle the whole thing
    df = df.sample(frac=1, random_state=SEED).reset_index(drop=True)

    return df


TICKET_TEMPLATES = {
    "Hardware": [
        "Nefunguje mi {hw}. Můžete se na to podívat?",
        "{hw} přestal/a fungovat dnes ráno. Potřebuji to vyřešit co nejdřív.",
        "Prosím o výměnu {hw} — je už za zenitem.",
    ],
    "Software": [
        "Nejde mi nainstalovat {sw}. Vyhazuje to chybu.",
        "Po updatu {sw} se mi rozbilo {feature}.",
        "Mohli byste mi prosím přidat licenci na {sw}?",
    ],
    "Účet a přístup": [
        "Zapomněl/a jsem heslo do {system}.",
        "Nemám přístup ke sdílenému disku týmu.",
        "Můj VPN profil nefunguje od pondělí.",
    ],
    "Ostatní": [
        "Nefunguje kávovar v 3. patře.",
        "Někdo mi v ledničce snědl oběd. Co s tím?",
        "Klimatizace v open-spacu zase netáhne.",
    ],
}
HW = ["notebook", "monitor", "myš", "klávesnice", "dock", "headset"]
SW = ["VS Code", "Slack", "Excel", "Jira", "Figma"]
FEATURE = ["sdílení obrazovky", "notifikace", "VPN integrace", "SSO login"]
SYSTEM = ["Jira", "Confluence", "interní portál", "Slack"]


def build_tickets(universe: pd.DataFrame) -> pd.DataFrame:
    rows = []
    ids = universe["employee_id"].tolist()
    start = pd.Timestamp("2024-01-01")
    for ticket_id in range(1, 5001):
        reporter = int(RNG.choice(ids))
        cat = str(RNG.choice(list(TICKET_TEMPLATES.keys()), p=[0.35, 0.35, 0.20, 0.10]))
        tmpl = str(RNG.choice(TICKET_TEMPLATES[cat]))
        popis = tmpl.format(
            hw=RNG.choice(HW), sw=RNG.choice(SW),
            feature=RNG.choice(FEATURE), system=RNG.choice(SYSTEM),
        )
        when = start + pd.Timedelta(days=int(RNG.integers(0, 500)),
                                    hours=int(RNG.integers(8, 18)),
                                    minutes=int(RNG.integers(0, 60)))
        rows.append({
            "ticket_id": ticket_id,
            "reporter_id": reporter,
            "datum": when.strftime("%Y-%m-%d %H:%M:%S"),
            "kategorie": cat,
            "priorita": str(RNG.choice(["P1", "P2", "P3", "P4"], p=[0.05, 0.20, 0.50, 0.25])),
            "status": str(RNG.choice(["open", "pending", "closed"], p=[0.15, 0.20, 0.65])),
            "popis": popis,
        })
    return pd.DataFrame(rows)


def build_org_chart(universe: pd.DataFrame) -> pd.DataFrame:
    """One row per employee per reporting line.

    ~80% of employees have 1 row; ~20% have 2–3 rows; cycles and orphan
    manager_ids are planted.
    """
    rows = []
    universe = universe.copy()
    universe["plat"] = pd.to_numeric(universe["plat"], errors="coerce").fillna(0)

    # Heuristic manager picker: choose someone in the same department with higher salary
    by_dept = {dept: sub for dept, sub in universe.groupby("oddeleni")}
    ids = universe["employee_id"].tolist()
    for _, emp in universe.iterrows():
        dept_peers = by_dept.get(emp["oddeleni"])
        candidates = dept_peers[dept_peers["plat"] > emp["plat"]] if dept_peers is not None else None
        if candidates is None or len(candidates) == 0:
            mgr = None  # department head
        else:
            mgr = int(candidates.sample(1, random_state=int(emp["employee_id"])).iloc[0]["employee_id"])
        hire = str(emp["datum_nastupu"])
        rows.append({
            "employee_id": int(emp["employee_id"]),
            "manager_id": mgr,
            "platnost_od": hire if hire and hire != "nan" else "2020-01-01",
        })
        # ~20% get a manager change
        if RNG.random() < 0.20 and mgr is not None:
            new_mgr = int(RNG.choice(ids))
            rows.append({
                "employee_id": int(emp["employee_id"]),
                "manager_id": new_mgr,
                "platnost_od": "2024-01-15",
            })

    df = pd.DataFrame(rows)

    # Plant 3 cycles: pick 3 pairs and swap
    for _ in range(3):
        pair = RNG.choice(df["employee_id"].unique(), size=2, replace=False)
        a, b = int(pair[0]), int(pair[1])
        a_latest = df[df["employee_id"] == a].tail(1).index
        b_latest = df[df["employee_id"] == b].tail(1).index
        df.loc[a_latest, "manager_id"] = b
        df.loc[b_latest, "manager_id"] = a

    # Plant 15 orphan manager_ids (point at IDs not in universe)
    orphan_idx = RNG.choice(df.index, size=15, replace=False)
    df.loc[orphan_idx, "manager_id"] = RNG.integers(9000, 9999, size=15)

    return df.sample(frac=1, random_state=SEED).reset_index(drop=True)


# ── Review text generation ────────────────────────────────────────

REVIEW_TEMPLATES_POSITIVE = [
    "{name} je velmi {adj_pos} a {praise}. {recommendation_pos}",
    "{name} odvádí skvělou práci{area}. {praise}. {recommendation_pos}",
    "S prací {name_gen} jsem velmi spokojený/á. {praise}. {recommendation_pos}",
    "{name} je klíčový/á člen/ka týmu. {praise}. {note_pos}",
]

REVIEW_TEMPLATES_MIXED = [
    "{name} je {adj_pos}, ale {weakness}. {recommendation_mix}",
    "{name} odvádí solidní práci, ale {weakness}. {recommendation_mix}",
    "S prací {name_gen} jsem celkově spokojený/á, nicméně {weakness}. {recommendation_mix}",
    "{name} má {strength_detail}, ale {weakness}. {recommendation_mix}",
]

REVIEW_TEMPLATES_NEGATIVE = [
    "{name} má problémy s {problem}. {neg_detail}. {recommendation_neg}",
    "Výkon {name_gen} je pod očekáváním. {neg_detail}. {recommendation_neg}",
    "U {name_gen} vidím opakované problémy s {problem}. {recommendation_neg}",
]

ADJ_POS = [
    "spolehlivý/á", "pracovitý/á", "kreativní", "pečlivý/á",
    "proaktivní", "samostatný/á", "komunikativní", "flexibilní",
]
PRAISE = [
    "klienti si ho/ji velmi pochvalují",
    "kolegové ho/ji respektují",
    "vždy dodržuje termíny",
    "přináší nové nápady do týmu",
    "jeho/její práce je vždy precizní",
    "zvládá i náročné situace s klidem",
    "je ochoten/ochotna pomoci kolegům",
    "dobře komunikuje s ostatními odděleními",
]
RECOMMENDATION_POS = [
    "Doporučuji zvážit povýšení.",
    "Navrhuji navýšení platu.",
    "Měli bychom mu/jí nabídnout více odpovědnosti.",
    "Výborný/á kandidát/ka na vedení projektu.",
    "",
]
NOTE_POS = [
    "Pokračovat v současném nastavení.",
    "Zvážit rozšíření kompetencí.",
    "Výborný přínos pro tým.",
]
STRENGTH_DETAIL = [
    "silné analytické schopnosti",
    "výborný přehled o produktech",
    "dobré vztahy s klienty",
    "technické znalosti na vysoké úrovni",
]
WEAKNESS = [
    "někdy nestíhá deadliny",
    "komunikace s týmem by mohla být lepší",
    "občas se ztrácí v detailech a zpomaluje celkový progres",
    "prezentační dovednosti potřebují zlepšení",
    "je vidět, že ho/ji současná pozice už moc nebaví",
    "má tendenci pracovat izolovaně",
    "občas dělá chyby v dokumentaci",
    "mohl/a by být aktivnější na poradách",
]
RECOMMENDATION_MIX = [
    "Doporučuji školení v oblasti time managementu.",
    "Navrhuji mentoring se seniornějším kolegou.",
    "Měli bychom zvážit přesun do jiného týmu.",
    "Doporučuji zpětnou vazbu od kolegů (360°).",
    "Zvážit rozvojový plán na příští kvartál.",
]
PROBLEM = [
    "dodržováním termínů",
    "kvalitou výstupů",
    "docházkou",
    "komunikací v týmu",
    "plněním KPI",
    "adaptací na nové procesy",
]
NEG_DETAIL = [
    "Opakovaně jsme řešili stížnosti od kolegů",
    "Výstupy vyžadují časté přepracování",
    "Několikrát nedodržel/a dohodnuté termíny",
    "Spolupráce s ostatními odděleními vázne",
]
RECOMMENDATION_NEG = [
    "Nutné zlepšení do konce zkušební doby.",
    "Doporučuji formální rozvojový plán s měsíčním hodnocením.",
    "Pokud se situace nezlepší, zvážit ukončení spolupráce.",
    "Navrhuji intenzivní koučink.",
]
AREAS = [
    "", " v oblasti analýzy dat", " na klientských projektech",
    " při správě interních systémů", " v komunikaci se zákazníky",
]


def _fill_review_template(template: str, name: str) -> str:
    """Fill a review template with random slot values."""
    # Genitive form — crude but good enough for templates
    if name.endswith("a"):
        name_gen = name[:-1] + "y"
    elif name.endswith("k"):
        name_gen = name + "a"
    else:
        name_gen = name + "a"

    replacements = {
        "{name}": name,
        "{name_gen}": name_gen,
        "{adj_pos}": RNG.choice(ADJ_POS),
        "{praise}": RNG.choice(PRAISE),
        "{recommendation_pos}": RNG.choice(RECOMMENDATION_POS),
        "{note_pos}": RNG.choice(NOTE_POS),
        "{strength_detail}": RNG.choice(STRENGTH_DETAIL),
        "{weakness}": RNG.choice(WEAKNESS),
        "{recommendation_mix}": RNG.choice(RECOMMENDATION_MIX),
        "{problem}": RNG.choice(PROBLEM),
        "{neg_detail}": RNG.choice(NEG_DETAIL),
        "{recommendation_neg}": RNG.choice(RECOMMENDATION_NEG),
        "{area}": RNG.choice(AREAS),
    }
    result = template
    for key, val in replacements.items():
        result = result.replace(key, val)
    return result


def build_reviews(df: pd.DataFrame) -> pd.DataFrame:
    """Generate ~150 performance reviews for a subset of employees."""
    # Pick ~150 employees (spread across departments)
    review_rows = []
    for dept in DEPARTMENTS:
        dept_mask = df["oddeleni"].isin(
            [dept, dept.lower(), "Mktg"] if dept == "Marketing" else [dept]
        )
        dept_employees = df[dept_mask]
        n_reviews = max(5, int(len(dept_employees) * 0.16))
        selected = dept_employees.sample(
            n=min(n_reviews, len(dept_employees)), random_state=SEED
        )

        for _, emp in selected.iterrows():
            perf = emp["hodnoceni_vykonu"]
            name = emp["jmeno"]

            # Choose template type based on performance score
            if perf >= 4:
                template = RNG.choice(REVIEW_TEMPLATES_POSITIVE)
            elif perf >= 3:
                template = RNG.choice(REVIEW_TEMPLATES_MIXED)
            else:
                template = RNG.choice(REVIEW_TEMPLATES_NEGATIVE)

            # Planted signal: ~15% of high-perf employees get mixed/negative review
            if perf >= 4 and RNG.random() < 0.15:
                template = RNG.choice(REVIEW_TEMPLATES_MIXED)

            review_text = _fill_review_template(template, name)

            # Review year — most recent full year of employment
            review_rows.append({
                "employee_id": emp["employee_id"],
                "rok": RNG.integers(2023, 2026),
                "review_text": review_text,
            })

    return pd.DataFrame(review_rows)


SARCASTIC_REVIEWS = [
    "Pan {prijmeni} je velmi *kreativní* s deadliny — vždy nás překvapí.",
    "Paní {prijmeni} přináší do týmu *unikátní* energii. Někdy příliš.",
    "S {jmeno_gen} {prijmeni_gen} je *radost* spolupracovat na složitých projektech.",
    "{jmeno} {prijmeni} má velmi *originální* přístup k procesům.",
    "{jmeno} je *neuvěřitelně* samostatný/á — kolegové se ho/ji téměř nedotknou.",
]


def apply_dirt_reviews(df: pd.DataFrame, main_df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    # 5 sarcastic reviews — overwrite 5 random review_text values.
    # The first 3 always use the *kreativní* template so the test signal is reliable.
    sarcastic_idx = RNG.choice(df.index, size=5, replace=False)
    for k, i in enumerate(sarcastic_idx):
        emp = main_df[main_df["employee_id"] == df.at[i, "employee_id"]]
        if len(emp) == 0:
            jmeno, prijmeni = "Jan", "Novák"
        else:
            jmeno, prijmeni = str(emp.iloc[0]["jmeno"]), str(emp.iloc[0]["prijmeni"])
        # First 3 entries guaranteed to contain *kreativní*; last 2 are random
        if k < 3:
            tmpl = SARCASTIC_REVIEWS[0]
        else:
            tmpl = str(RNG.choice(SARCASTIC_REVIEWS[1:]))
        df.at[i, "review_text"] = tmpl.format(
            jmeno=jmeno, prijmeni=prijmeni,
            jmeno_gen=jmeno, prijmeni_gen=prijmeni,
        )

    # 3 wrong-person reviews — keep employee_id but replace the name in the text
    other = main_df.sample(3, random_state=SEED)
    wrong_idx = RNG.choice(df.index, size=3, replace=False)
    for k, i in enumerate(wrong_idx):
        ghost = other.iloc[k]
        df.at[i, "review_text"] = (
            f"{ghost['jmeno']} {ghost['prijmeni']} odvádí solidní práci, "
            f"ale rezervy jsou v komunikaci s týmem."
        )
    return df


# ── Exit interview text generation ────────────────────────────────

EXIT_TEMPLATES = {
    "salary": [
        "Odcházím hlavně kvůli platu. Po {years} letech jsem pořád na podobné úrovni a jinde mi nabídli výrazně víc. {note}",
        "Finanční ohodnocení neodpovídá mé práci. {note} Jinak nemám výhrady k týmu.",
        "Dostal/a jsem nabídku s o {diff}% vyšším platem. Tady nebylo na vyjednávání prostoru. {note}",
    ],
    "growth": [
        "Cítím, že jsem se tu přestal/a rozvíjet. Po {years} letech dělám pořád to samé. {note}",
        "Chybí mi možnost kariérního růstu. Chtěl/a bych se posunout do vedoucí pozice, ale tady to není reálné. {note}",
        "Odcházím, protože jsem dostal/a příležitost pracovat na zajímavějších projektech. {note}",
    ],
    "work_life_balance": [
        "Přecházím na pozici s možností home office. Tady to prostě nešlo skloubit s rodinou. {note}",
        "Potřebuji flexibilnější pracovní dobu. {note} Jinak jsem tu byl/a spokojený/á.",
        "Dojíždění mi zabírá příliš času. Našel/la jsem práci blíž k domovu. {note}",
    ],
    "management": [
        "Neshodl/a jsem se s vedením oddělení. {note} Komunikace shora dolů nefungovala.",
        "Změna vedení v posledním roce hodně ovlivnila atmosféru v týmu. {note}",
        "Chybí mi zpětná vazba a jasné směřování od managementu. {note}",
    ],
    "other": [
        "Stěhuji se do jiného města, takže dojíždění není možné. {note}",
        "Vracím se ke studiu na plný úvazek. {note} Rád/a bych se sem jednou vrátil/a.",
        "Dostal/a jsem příležitost v zahraničí, kterou nechci propásnout. {note}",
    ],
}

EXIT_NOTES_POS = [
    "Jinak super kolektiv.",
    "Děkuji za příležitosti, které jsem tu dostal/a.",
    "Kolegům budu děkovat za spolupráci.",
    "Bylo to tu fajn, jen ty podmínky už nestačí.",
    "",
]
EXIT_NOTES_NEG = [
    "Upřímně, moc mi to tu chybět nebude.",
    "Doufám, že se situace zlepší pro ostatní kolegy.",
    "Několik kolegů uvažuje o odchodu ze stejných důvodů.",
    "",
]


def build_exit_interviews(df: pd.DataFrame) -> pd.DataFrame:
    """Generate ~70 exit interview notes, skewed toward Podpora."""
    from datetime import timedelta

    exit_rows = []

    # Podpora: ~40 exits, others: ~30 total
    dept_exit_counts = {
        "Podpora": 40,
        "Marketing": 6,
        "Obchod": 8,
        "HR": 4,
        "Finance": 4,
        "Vývoj": 8,
    }

    # Reason distribution per department
    dept_reason_weights = {
        "Podpora": {"salary": 0.40, "growth": 0.30, "work_life_balance": 0.15,
                     "management": 0.10, "other": 0.05},
        "Marketing": {"salary": 0.20, "growth": 0.30, "work_life_balance": 0.20,
                       "management": 0.15, "other": 0.15},
        "Obchod": {"salary": 0.25, "growth": 0.20, "work_life_balance": 0.15,
                    "management": 0.20, "other": 0.20},
        "HR": {"salary": 0.20, "growth": 0.25, "work_life_balance": 0.25,
               "management": 0.15, "other": 0.15},
        "Finance": {"salary": 0.15, "growth": 0.20, "work_life_balance": 0.20,
                     "management": 0.15, "other": 0.30},
        "Vývoj": {"salary": 0.30, "growth": 0.25, "work_life_balance": 0.15,
                   "management": 0.15, "other": 0.15},
    }

    for dept, n_exits in dept_exit_counts.items():
        dept_mask = df["oddeleni"].isin(
            [dept, dept.lower(), "Mktg"] if dept == "Marketing" else [dept]
        )
        dept_employees = df[dept_mask]
        if len(dept_employees) < n_exits:
            n_exits = len(dept_employees)

        import hashlib as _hashlib
        dept_seed = SEED + int(_hashlib.sha256(dept.encode()).hexdigest(), 16) % 1000
        selected = dept_employees.sample(n=n_exits, random_state=dept_seed)

        reasons = list(dept_reason_weights[dept].keys())
        weights = list(dept_reason_weights[dept].values())

        for _, emp in selected.iterrows():
            reason = RNG.choice(reasons, p=weights)
            template = RNG.choice(EXIT_TEMPLATES[reason])

            # Compute years roughly from datum_nastupu
            try:
                from datetime import datetime

                if "." in str(emp["datum_nastupu"]):
                    hire = datetime.strptime(str(emp["datum_nastupu"]), "%d.%m.%Y")
                else:
                    hire = datetime.strptime(str(emp["datum_nastupu"]), "%Y-%m-%d")
                years = round((datetime(2025, 6, 1) - hire).days / 365, 1)
            except (ValueError, TypeError):
                years = 2.0

            # Pick note based on whether exit is bitter or amicable
            is_bitter = reason in ("management", "salary") and RNG.random() < 0.4
            note = RNG.choice(EXIT_NOTES_NEG if is_bitter else EXIT_NOTES_POS)

            text = template.format(
                years=years,
                diff=RNG.integers(15, 40),
                note=note,
            ).strip()

            # Exit date — sometime in 2024-2025
            from datetime import date

            exit_date = date(2024, 1, 1) + timedelta(days=int(RNG.integers(0, 540)))

            exit_rows.append({
                "employee_id": emp["employee_id"],
                "datum_odchodu": exit_date.strftime("%Y-%m-%d"),
                "interview_text": text,
            })

    return pd.DataFrame(exit_rows)


EN_FRAGMENTS = [
    " Hlavní důvod: better compensation elsewhere. Tým byl ok.",
    " management was nice, ale firma jako celek nepostupuje.",
    " I'm looking for better opportunity. Děkuji za vše.",
    " Tým, team dynamics, byl super, ale management nereaguje.",
    " Going for a senior role elsewhere — tady už nebyl prostor.",
]
CONTRADICTION = (
    "Platově jsem byl celkem spokojen, kolegové fajn. Odcházím hlavně kvůli "
    "tomu, že peníze už nejsou dostatečné a HR nedokázalo nic nabídnout."
)


def apply_dirt_exit_interviews(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    en_idx = RNG.choice(df.index, size=5, replace=False)
    for k, i in enumerate(en_idx):
        df.at[i, "interview_text"] = str(df.at[i, "interview_text"]) + EN_FRAGMENTS[k % len(EN_FRAGMENTS)]
    # 1 self-contradiction — must not overwrite one of the EN_FRAGMENT rows
    remaining = [i for i in df.index if i not in set(en_idx)]
    contra_idx = RNG.choice(remaining, size=1)
    df.loc[contra_idx, "interview_text"] = CONTRADICTION
    return df


def build_payroll_xlsx(main_df: pd.DataFrame) -> pd.DataFrame:
    """Return the data frame and *also* writes it directly to xlsx in main().

    For testability the function only computes the payroll rows; the xlsx
    write (with merged headers + totals row + extra sheets) lives in
    `write_payroll_xlsx` below.
    """
    rows = []
    for _, emp in main_df.iterrows():
        plat = pd.to_numeric(emp.get("plat"), errors="coerce")
        if pd.isna(plat):
            plat = 50_000.0
        bonus = int(RNG.integers(0, 15_000))
        rows.append({
            "os_cislo": int(emp["employee_id"]),
            "jmeno_prijmeni": f"{str(emp['jmeno']).strip()} {str(emp['prijmeni']).strip()}",
            "oddeleni": emp["oddeleni"],
            "mzda_brutto": float(plat),
            "bonus": bonus,
            "mzda_celkem": float(plat) + bonus,
            "aktivni": "ano",
        })
    return pd.DataFrame(rows)


def apply_dirt_payroll(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    # 8% of mzda_brutto off by 1-15%
    off_idx = RNG.choice(df.index, size=int(len(df) * 0.08), replace=False)
    for i in off_idx:
        delta = 1 + RNG.uniform(-0.15, 0.15)
        df.at[i, "mzda_brutto"] = round(df.at[i, "mzda_brutto"] * delta)
        df.at[i, "mzda_celkem"] = df.at[i, "mzda_brutto"] + df.at[i, "bonus"]

    # 30 rows where mzda_celkem != mzda_brutto + bonus
    bad_arith = RNG.choice(df.index, size=30, replace=False)
    for i in bad_arith:
        df.at[i, "mzda_celkem"] = df.at[i, "mzda_brutto"] + df.at[i, "bonus"] - int(RNG.integers(500, 3000))

    # 4 EUR-not-CZK landmines
    eur_idx = RNG.choice(df.index, size=4, replace=False)
    for i in eur_idx:
        df.at[i, "mzda_brutto"] = round(df.at[i, "mzda_brutto"] / 25)
        df.at[i, "bonus"] = round(df.at[i, "bonus"] / 25)
        df.at[i, "mzda_celkem"] = df.at[i, "mzda_brutto"] + df.at[i, "bonus"]

    # 5 English-language oddeleni values
    en_map = {"Vývoj": "Engineering", "Podpora": "Support", "Marketing": "Marketing",
              "Obchod": "Sales", "Finance": "Finance", "HR": "HR"}
    en_idx = RNG.choice(df.index, size=5, replace=False)
    for i in en_idx:
        cs = df.at[i, "oddeleni"]
        df.at[i, "oddeleni"] = en_map.get(cs, cs)

    # 3 plausible-wrong departments (rotate cs values)
    wrong_idx = RNG.choice(df.index, size=3, replace=False)
    rotation = ["Marketing", "Vývoj", "Obchod"]
    for k, i in enumerate(wrong_idx):
        df.at[i, "oddeleni"] = rotation[k]

    # `aktivni` mixed truthiness
    n = len(df)
    forms = ["ano", "Ano", "ANO", "y", "1", "true", ""]
    for k, form in enumerate(forms):
        m = max(5, n // (len(forms) * 6))
        sample_idx = RNG.choice(df.index, size=m, replace=False)
        df.loc[sample_idx, "aktivni"] = form

    return df


def write_payroll_xlsx(df: pd.DataFrame, path: Path) -> None:
    """Write payroll DF with 3 merged-header rows, totals row, and extra sheets.

    The xlsx is written with a fixed creation/modification timestamp so that
    repeated runs produce byte-identical output (openpyxl otherwise injects
    the current wall-clock time into docProps/core.xml on every save).
    """
    import io as _io
    import re as _re
    import zipfile as _zf
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Mzdy Q3 2025"
    # 3 merged-header artifact rows
    ws.append(["Mzdy zaměstnanců — Q3 2025", None, None, None, None, None, None])
    ws.append([None, None, None, None, None, None, None])
    ws.append(["DataCorp s.r.o. — interní použití", None, None, None, None, None, None])
    # Header row
    ws.append(list(df.columns))
    for r in df.itertuples(index=False):
        ws.append(list(r))
    # CELKEM totals row
    totals = ["CELKEM", None, None,
              float(df["mzda_brutto"].sum()), float(df["bonus"].sum()),
              float(df["mzda_celkem"].sum()), None]
    ws.append(totals)
    wb.create_sheet("List2")
    wb.create_sheet("List3")
    # Save to a buffer first, then rewrite with a fixed timestamp
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    out = _io.BytesIO()
    _FIXED_TS = "2025-10-01T00:00:00Z"
    _FIXED_DATE = (2025, 10, 1, 0, 0, 0)  # fixed DOS timestamp for every entry
    with _zf.ZipFile(buf, "r") as zin, _zf.ZipFile(out, "w", compression=_zf.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "docProps/core.xml":
                data = _re.sub(
                    rb"<dcterms:created[^>]*>.*?</dcterms:created>",
                    f"<dcterms:created xsi:type=\"dcterms:W3CDTF\">{_FIXED_TS}</dcterms:created>".encode(),
                    data,
                )
                data = _re.sub(
                    rb"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                    f"<dcterms:modified xsi:type=\"dcterms:W3CDTF\">{_FIXED_TS}</dcterms:modified>".encode(),
                    data,
                )
            item.date_time = _FIXED_DATE
            zout.writestr(item, data)
    path.write_bytes(out.getvalue())


def apply_dirt_tickets(df: pd.DataFrame, universe: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    # Taxonomy mismatch — replace ~7% of "Hardware" with variants
    hw_mask = df["kategorie"] == "Hardware"
    hw_idx = df[hw_mask].index.tolist()
    variants = ["HW", "hardware-other", "hw / sw", "HARDWARE", "Hardware "]
    chunk = max(1, len(hw_idx) // (len(variants) * 7))
    cursor = 0
    for v in variants:
        slice_idx = hw_idx[cursor: cursor + chunk]
        df.loc[slice_idx, "kategorie"] = v
        cursor += chunk

    # Priorita mismatch — bump 40% of P3/P4 with trivial popis to P1
    coffee_mask = df["popis"].str.contains("kávovar|lednič|klimatizac", na=False)
    coffee_idx = df[coffee_mask].sample(frac=0.4, random_state=SEED).index
    df.loc[coffee_idx, "priorita"] = "P1"

    # 50 invalid reporter_ids
    valid_ids = set(universe["employee_id"])
    invalid_idx = RNG.choice(df.index, size=50, replace=False)
    df.loc[invalid_idx, "reporter_id"] = RNG.integers(9000, 9999, size=50)

    # 10 trailing-punctuation / leading-whitespace IDs
    punct_idx = RNG.choice(df.index, size=10, replace=False)
    for k, i in enumerate(punct_idx):
        rid = str(df.at[i, "reporter_id"])
        df.at[i, "reporter_id"] = f"{rid}." if k % 2 == 0 else f" {rid}"

    return df


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print("Building employee universe...")
    universe = build_main_df()
    universe = tag_departures(universe)
    print(f"  Universe: {len(universe)} ({universe['_departed'].sum()} departed)")

    main_df = universe[~universe["_departed"]].drop(columns=["_departed"]).copy()
    main_df_dirty = apply_dirt_main(main_df)
    universe_for_sides = universe.drop(columns=["_departed"]).copy()

    print("Building side-tables...")
    salary_history = apply_dirt_salary_history(
        build_salary_history(universe_for_sides), main_df_dirty,
    )
    org = build_org_chart(universe_for_sides)
    tickets = apply_dirt_tickets(build_tickets(universe_for_sides), universe_for_sides)
    payroll = apply_dirt_payroll(build_payroll_xlsx(main_df_dirty))
    reviews = apply_dirt_reviews(build_reviews(universe_for_sides), main_df_dirty)
    exits = apply_dirt_exit_interviews(build_exit_interviews(universe_for_sides))

    print("Writing files...")
    main_df_dirty.to_csv(OUTPUT_DIR / "datacorp.csv", index=False)
    reviews.to_csv(OUTPUT_DIR / "datacorp_reviews.csv", index=False)
    exits.to_csv(OUTPUT_DIR / "datacorp_exit_interviews.csv", index=False)
    salary_history.to_csv(OUTPUT_DIR / "datacorp_salary_history.csv", index=False)
    org.to_csv(OUTPUT_DIR / "datacorp_org_chart.csv", index=False)
    tickets.to_csv(OUTPUT_DIR / "datacorp_tickets.csv", index=False)
    write_payroll_xlsx(payroll, OUTPUT_DIR / "datacorp_payroll_q3.xlsx")

    # Smoke checks
    assert 990 <= len(main_df_dirty) <= 1030
    assert 2500 <= len(salary_history) <= 4000
    assert 1000 <= len(org) <= 1400
    assert 4500 <= len(tickets) <= 5500
    print("Done!")


if __name__ == "__main__":
    main()
